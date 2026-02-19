#!/usr/bin/env python3
"""
core_sms_sender.py

Core logic for sending bulk SMS messages. This module is intended to be used by both
the command-line interface and the future GUI application.
"""
from __future__ import annotations
import csv
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Any, Iterable, List, Tuple, Optional
import requests
import openpyxl
from tkinter import messagebox

# ----------------- Helpers -----------------

class SafeDict(dict):
    def __missing__(self, key):
        return ""

def normalize_phone(phone: Optional[str], country_prefix: Optional[str]=None, ensure_plus: bool=False) -> str:
    if phone is None:
        return ""
    s = str(phone).strip()
    for ch in (" ", "-", "(", ")", ".", "\u200e", "\u200f"):
        s = s.replace(ch, "")
    if s == "":
        return ""
    if s.startswith("00"):
        s = "+" + s[2:]
    if not s.startswith("+") and country_prefix:
        cp_digits = country_prefix.lstrip("+")
        if not s.startswith(cp_digits):
            s = s.lstrip("0")
            s = country_prefix + s
    if ensure_plus and not s.startswith("+"):
        s = "+" + s
    return s

def build_message(template: str, row: Dict[str, Any]) -> str:
    clean = {k: ("" if v is None else str(v)) for k, v in row.items()}
    for k, v in list(clean.items()):
        clean[f"{k}_upper"] = v.upper()
        clean[f"{k}_lower"] = v.lower()
    return template.format_map(SafeDict(clean))

def send_to_gateway(session: requests.Session, url: str, to: str, message: str, auth: Optional[str], timeout: float
                   ) -> Tuple[bool, Optional[int], str]:
    headers = {"Content-Type": "application/json"}
    if auth:
        headers["Authorization"] = auth
    payload = {"to": to, "message": message}
    try:
        resp = session.post(url, json=payload, headers=headers, timeout=timeout)
        text = resp.text if resp is not None else ""
        return (resp.ok if resp is not None else False, resp.status_code if resp is not None else None, text)
    except requests.RequestException as e:
        return (False, None, str(e))

# ----------------- File reading utilities -----------------

def get_file_headers(path: Path) -> List[str]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
    elif path.suffix.lower() in [".xlsx", ".xls"]:
        workbook = openpyxl.load_workbook(path, read_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
    else:
        headers = []
    return headers

def get_row_count(path: Path) -> Optional[int]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return sum(1 for row in f) - 1 # Exclude header
    elif path.suffix.lower() in [".xlsx", ".xls"]:
        try:
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook.active
            return sheet.max_row - 1 # Exclude header
        except Exception:
            return None
    return None


def iter_rows(path: Path, start_row: int) -> Iterable[Dict[str, str]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            for i, raw_row in enumerate(reader, 1):
                if i < start_row:
                    continue
                if len(raw_row) < len(headers):
                    raw_row += [""] * (len(headers) - len(raw_row))
                row = {headers[i]: raw_row[i] for i in range(len(headers))}
                if "phone" not in {h.lower() for h in headers}:
                    if len(raw_row) >= 3 and "phone" not in row:
                        row.setdefault("phone", raw_row[2])
                yield row
    elif path.suffix.lower() in [".xlsx", ".xls"]:
        workbook = openpyxl.load_workbook(path, read_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        for i, sheet_row in enumerate(sheet.iter_rows(min_row=2), 1):
            if i < start_row:
                continue
            raw_row = [cell.value for cell in sheet_row]
            if len(raw_row) < len(headers):
                raw_row += [""] * (len(headers) - len(raw_row))
            row = {headers[i]: raw_row[i] for i in range(len(headers))}
            if "phone" not in {h.lower() for h in headers}:
                if len(raw_row) >= 3 and "phone" not in row:
                    row.setdefault("phone", raw_row[2])
            yield row

# ----------------- Main sending logic -----------------

class SMSSender:
    def __init__(self, config, db_conn=None):
        self.config = config
        self.db_conn = db_conn
        self.session = requests.Session()
        self.continue_sending = True
        self.sent_count = 0

    def get_daily_sent_count(self):
        if not self.db_conn:
            return 0
        with self.db_conn.cursor() as cur:
            cur.execute("""
                SELECT COUNT(*) FROM messages 
                WHERE gateway_id = %s AND DATE(sent_at) = CURRENT_DATE
            """, (self.config["gateway_id"],))
            return cur.fetchone()[0]
            
    def was_sent_today(self, phone_number: str) -> bool:
        if not self.db_conn:
            return False
        with self.db_conn.cursor() as cur:
            cur.execute("""
                SELECT 1 FROM messages 
                WHERE phone_number = %s AND status = 'success' AND DATE(sent_at) = CURRENT_DATE
            """, (phone_number,))
            return cur.fetchone() is not None

    def log_to_db_batch(self, logs):
        if self.db_conn and logs:
            with self.db_conn.cursor() as cur:
                cur.executemany(
                    "INSERT INTO messages (gateway_id, phone_number, message, status) VALUES (%s, %s, %s, %s)",
                    logs
                )
                self.db_conn.commit()

    def send_messages(self, contact_file: Path, template: str, total_rows: Optional[int], dry_run: bool = False, on_progress=None, on_complete=None):
        headers = get_file_headers(contact_file)
        out_fields = list(headers) + ["phone_normalized", "message", "attempts", "success", "status_code", "response", "sent_at", "sent_status"]
        output_file = self.config.get("output_file", "sms_send_log.csv")
        
        self.sent_count = self.get_daily_sent_count()

        log_batch = []
        batch_size = 100
        
        with open(output_file, "w", encoding="utf-8", newline="") as out_f:
            writer = csv.DictWriter(out_f, fieldnames=out_fields)
            writer.writeheader()

            total_processed = 0
            
            for idx, row in enumerate(iter_rows(contact_file, self.config.get("start_row", 1)), start=self.config.get("start_row", 1)):
                if not self.continue_sending:
                    if on_progress: on_progress("Sending halted by user.", total_processed)
                    break

                if self.sent_count >= 1000:
                    if not messagebox.askyesno("Rate Limit Reached", f"{self.sent_count} messages have been sent with this gateway today. Do you want to continue?"):
                        self.continue_sending = False
                        continue
                
                if self.config.get("limit") and total_processed >= self.config.get("limit"):
                    break
                total_processed += 1
                
                raw_phone = row.get("phone", "") or row.get("Phone", "") or ""
                phone_norm = normalize_phone(raw_phone, country_prefix=self.config.get("country_prefix"), ensure_plus=self.config.get("ensure_plus"))
                
                if self.config.get("skip_duplicates") and self.was_sent_today(phone_norm):
                    if on_progress:
                        on_progress(f"Skipping row {idx} (already sent today): {phone_norm}", total_processed)
                    continue

                message = build_message(template, row)

                log_row = {h: row.get(h, "") for h in headers}
                log_row.update({"phone_normalized": phone_norm, "message": message, "attempts": 0, "success": False, "status_code": "", "response": "", "sent_at": "", "sent_status": 0})
                
                if dry_run:
                    if on_progress:
                        on_progress(f"[DRY] Row {idx} -> To: {phone_norm}", total_processed)
                    writer.writerow(log_row)
                    continue

                attempt = 0
                success = False
                last_status = None
                last_resp = ""
                while attempt <= self.config.get("retries", 2):
                    attempt += 1
                    log_row["attempts"] = attempt
                    ok, status_code, resp_text = send_to_gateway(self.session, self.config["gateway_url"], phone_norm, message, auth=self.config.get("auth"), timeout=self.config.get("timeout", 10.0))
                    last_status = status_code
                    last_resp = resp_text
                    if ok:
                        success = True
                        break
                    backoff = min(5.0, 0.5 * (2 ** (attempt - 1)))
                    time.sleep(backoff)
                
                log_row["success"] = success
                log_row["status_code"] = "" if last_status is None else str(last_status)
                log_row["response"] = last_resp
                log_row["sent_at"] = datetime.utcnow().isoformat() + "Z"
                log_row["sent_status"] = 1 if success else 0
                writer.writerow(log_row)
                
                if success:
                    self.sent_count += 1
                
                log_batch.append((self.config["gateway_id"], phone_norm, message, "success" if success else "failed"))
                if len(log_batch) >= batch_size:
                    self.log_to_db_batch(log_batch)
                    log_batch = []

                if total_processed % 100 == 0:
                    if on_progress:
                        on_progress(f"Processed {total_processed} rows...", total_processed)
                
                time.sleep(self.config.get("delay", 0.2))

            # Log any remaining messages
            self.log_to_db_batch(log_batch)

        if on_complete:
            on_complete(f"Processed {total_processed} rows. Log written to {output_file}")
        
        if self.db_conn:
            self.db_conn.close()
